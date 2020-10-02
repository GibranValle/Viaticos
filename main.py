"""
Proyecto final del curso: "Python en el ámbito científico"
IPN CIC 2020
REVISION FINAL: 08/06/2020

Gibran Valle
Revisión final: 13/03/2020

    Este proyecto tiene como finalidad agilizar la comprobación de viaticos
    que se erogan por los ingenieros de servicio en la empresa SEVIME.
    Está lleno de ideas ambiciosas y aunque la premisa no incluye algoritmos
    muy complejos, algunos bots ganaron cierto grado de dificultad.

Instrucciones:
    1)  El usuario debe ingresar un archivo .xls en el directorio raiz
    2)  El programa genera un nuevo archivo .xlsx con mejor formato
    3)  El usuario debe copiar los valores de este nuevo archivo .xlsx en la base de datos
    4)  El programa separa los gastos no deducibles, genera una plantilla formulario .pdf
        que se puede editar posteriormente con cualquier programa de lectura de PDF

    Notas:
    1)  El proyecto funciona correctamente utilizando las plantillas de vale estándar
        TODO: IMPLEMENTAR LA GENERACIÓN Y LLENADO AUTOMÁTICO DE UN VALE PERSONALIZADO
    2)  El proyecto necesita más tiempo para pulir el algoritmo final, el cual
        generará la plantilla formulario de reporte semanal y lo llenará automáticamente
    3)  Cuando todos los algoritmos estén listos, se empaquetará el programa para
        que los empleados de la empresas puedan utilizarlo en las PC de la oficina.

    VISIÓN DEL PROYECTO
    A)  Como mejoras pendientes que no podré implementar por falta de conocimiento de base de datos:
        TODO: EXPORTAR LOS VALORES A UNA BASE DE DATOS ACCESIBLE DESDE LA NUBE
    B)  Como proyecto paralelo se desea poder crear una app específica que logre implementar
        todos los algoritmos de este proyecto.
"""
import pandas as pd

from bots.botClasificador import clasificador as classify
from bots.botClasificador import tipoComprobante as tipo
from bots.botContador import saldoDiario as sd
from bots.botEstructurador import estructurarFecha as ef
from bots.botEstructurador import estructurarImporte as ei
from bots.botFormuladorVale import llenarVales as cv

pd.options.mode.chained_assignment = None  # default='warn'

from bots.botReportador import organizador as og

import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, NumberFormatDescriptor, Alignment
from openpyxl.styles import numbers, alignment
from openpyxl.utils import get_column_letter
import re

# TAG PARA PDF
usuario = "Gibran Valle"

diccionario_vale = {
    'Importe': '100.00',
    'ImporteLetra': '(Cien pesos 00 M.N)',
    'Concepto A': '',
    'Concepto B': '',
    'Concepto C': '',
    "Fecha": "",
    "Recibido": usuario
}

# columnas extraidas de la app "Mi presupuesto"
columns_app = [
    "fecha",
    "concepto",
    "detalles",
    "metodo_pago",
    "categoria",
    "subcategoria",
    "proveedor",
    "grupo",
    "cuenta",
    "importe",
]
# columnas para borrar
drop_cols = ["metodo_pago", "subcategoria", "cuenta", "categoria"]
# columnas ordenadas segun el reporte usado
final_columns = [
    "fecha",
    "mes",
    "dia",
    "semana",
    "orden",
    "concepto",
    "detalles",
    "proveedor",
    "importe",
    "saldo diario",
    "comprobante",
    "clasificacion",
    "grupo"
]
ingreso_col = [
    "semana",
    "fecha",
    "importe",
]
# columnas a usar para el reporte pdf
columns_report = [
    "FECHA",
    "CASETAS",
    "GASOLINA",
    "HOSPEDAJE",
    "PAQUETERIA",
    "REFA - HERRA",
    "ESTACIONAMIENTO",
    "OTROS",
    "ALIMENTO - TAXI 1",
    "ALIMENTO - TAXI 2",
    "ALIMENTO - TAXI 3",
    "ALIMENTO - TAXI 4",
    "ALIMENTO - TAXI 5",
    "ALIMENTO - TAXI 6",
    "ALIMENTO - TAXI 7",
    "ALIMENTO - TAXI 8",
    "ALIMENTO - TAXI 9"
]
# columnas extraidas del excel
columns_excel = [
    "fecha",
    "dia",
    "semana",
    "orden",
    "concepto",
    "detalles",
    "proveedor",
    "importe",
    "saldo diario",
    "comprobante",
    "clasificacion",
    "grupo"
]


# --------------- FUNCIONES DEL PROGRAMA ------------------------------------
def as_text(value):
    if value is None:
        return ""
    return str(value)


def aplicarFormatos():
    # estilo de borde
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")

    # color de celda
    my_red = openpyxl.styles.colors.Color(rgb='F2F2F2')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

    # elementos de formato
    normal = NamedStyle(name="normal")
    normal.font = Font(bold=False, size=11, name="verdana")
    normal.number_format = numbers.FORMAT_GENERAL
    normal.alignment.wrap_text = False
    # 'bottom', 'top', 'distributed', 'justify', 'center'
    normal.alignment.vertical = "center"
    normal.alignment.horizontal = "center"
    normal.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # elementos de formato
    bold = NamedStyle(name="bold")
    bold.font = Font(bold=True, size=11, name="verdana")
    bold.number_format = numbers.FORMAT_GENERAL
    bold.alignment.wrap_text = False
    # 'bottom', 'top', 'distributed', 'justify', 'center'
    bold.alignment.vertical = "center"
    bold.alignment.horizontal = "center"
    bold.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # formato de fecha
    fecha = NamedStyle(name="fecha")
    fecha.font = Font(bold=True, size=11, name="verdana")
    fecha.number_format = numbers.FORMAT_DATE_DMYSLASH
    fecha.alignment.wrap_text = False
    # 'bottom', 'top', 'distributed', 'justify', 'center'
    fecha.alignment.vertical = "center"
    fecha.alignment.horizontal = "center"
    fecha.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # formato de numero
    money = NamedStyle(name="money")
    money.font = Font(bold=False, size=11, name="verdana")
    money.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    money.alignment.wrap_text = False
    # 'bottom', 'top', 'distributed', 'justify', 'center'
    money.alignment.vertical = "center"
    money.alignment.horizontal = "center"
    money.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # iterar por todas las filas
    for i, row in enumerate(ws.iter_rows(ws.min_row, ws.max_row)):
        # print(i)
        ws.row_dimensions[i + 1].height = 40
        # print(row[0])
        # iterar en cada columna de la fila
        for cell in row:
            column_number = cell.column
            column_letter = get_column_letter(column_number)
            # print(column_letter)
            # print(cell)
            cell.style = normal
            cell.fill = my_fill

            if column_letter == "A":
                cell.style = fecha
                cell.fill = my_fill
            elif (column_letter == "H") or (column_letter == "I"):
                cell.style = money
                cell.fill = my_fill
            elif (column_letter == "L") or (column_letter == "J"):
                cell.style = bold
                cell.fill = my_fill
    return


def ajustarAncho():
    # iteracion por todas las columns para ajustar el ancho
    for columns in ws.columns:
        max_length = 0
        # print(column_cells)  imprime lista de todas las columnas
        column = columns[0]  # columna inicial
        column_number = column.column
        column_letter = get_column_letter(column_number)
        # print(columns[0])  #imprime solo la columna inicial
        # print(column)
        # print(column_letter)

        # iteracion por fila
        for i, cell in enumerate(columns):
            # print("celda: {},{}".format(str(column_number), i))
            l = len(str(cell.value))
            # print(l)
            if (l > max_length):
                max_length = l

            adjusted_width = (max_length + 2) * 1.3  # arbitrario
            ws.column_dimensions[column_letter].width = adjusted_width
    return


def crearFormula():
    # iteracion por filas para armar la formula
    for i, rows in enumerate(ws.rows):

        if i > 0:
            # SUMIFS FORMULAS
            # ws["M2"] = "=WEEKDAY(A2)"

            intervalo_semana = "C" + str(ws.min_row + 1) + ":C" + str(ws.max_row)
            intervalo_dia = "B" + str(ws.min_row + 1) + ":B" + str(ws.max_row)
            intervalo_suma = "H" + str(ws.min_row + 1) + ":H" + str(ws.max_row)

            # print(intervalo_dia)
            # print(intervalo_semana)
            # print(intervalo_suma)

            # =SUMAR.SI.CONJUNTO($I$25:$I$161;$C$25:$C$161;DIASEM(B34);$D$25:$D$161;NUM.DE.SEMANA(B34))
            # formula_excel = "=SUMAR.SI.CONJUNTO(" + intervalo_suma + "," + intervalo_dia + "," + "DIASEM(A" + str(
            # i + 1) + ")," + intervalo_semana + "," + "NUM.DE.SEMANA(A" + str(i + 1) + "))"
            formula = "=SUMAR.SI.CONJUNTO(" + intervalo_suma + "," + intervalo_dia + "," + "DIASEM(A" + str(
                i + 1) + ")," + intervalo_semana + "," + "NUM.DE.SEMANA(A" + str(i + 1) + "))"
            # print(formula)
            ws["I" + str(i + 1)] = formula
    return


# --------------- INICIO DEL PROGRAMA ------------------------------------


# --------------- INICIO DEL PROGRAMA ------------------------------------
# cargar archivo xls en el directorio
excel = "viaticos.xls"

# dataframe completo, si no hay fecha no es una columna valida
data = pd.read_excel(excel, header=0, names=columns_app)
data = data.dropna(subset=["fecha"])
print("Archivo inicial:\n")
print(data.head())

# --------------- LIMPIEZA DEL DATAFRAME ---------------------------------
# drop columnas
data.drop(drop_cols, axis=1, inplace=True)

# convertir la columna fecha a objeto de fecha y extraer componentes para procesamiento
data["fecha"], data["mes"], data["semana"], data["dia"] = ef(data["fecha"])

# quitar simbolos para estadistico (-$)
data["importe"] = ei(data.importe)

# --------------- ESTRUCTURACION DEL DATAFRAME ---------------------------
# clasificar los conceptos
data["clasificacion"] = classify(data["concepto"])
data["clasificacion"].fillna(value="otro", inplace=True)

# Crea columna de comprobante de pago, mapeamos los datos para recibir vale o factura
data["comprobante"] = tipo(data["concepto"], mapear=True)
# tratamos los demas valores como facturas
data["comprobante"].fillna(value="factura", inplace=True)

# --------------- PROCESAMIENTO DEL DATAFRAME ---------------------------
# Crear un dataframe con los vales a elaborar
vales = data[data.comprobante == "vale"]
# limpiar valores nan
vales["detalles"].fillna(value="", inplace=True)
cv(vales, usuario)

# TODO BUSCAR UNA MANERA DE AUTOMATIZAR ESTA PARTE
# crear columna para ordenar manualmente
data["orden"] = ""

# crear serie de importes positivos
positivos = data[data.importe > 0]
ingresos = pd.DataFrame(positivos)
ingresos.reset_index(drop=True, inplace=True)  # drop para no agregar la columna de indices viejos
# elegir unicamente las columnas ingreso_col
ingresos = ingresos[ingreso_col]
ingresos["detalles"] = ""
# elegir la primer columna de la serie detalles
ingresos["detalles"][0] = "INICIO DE MES"
print("\ndataframe ingresos:\n", ingresos)
# guardar como otro archivo
ingresos.to_excel("outputs/ingresos2020.xlsx", index=False)

# actualizar dataframe sin ingresos
data = data[data.importe < 0]
data.reset_index(drop=True, inplace=True)  # drop para no agregar la columna de indices viejos
data["saldo diario"] = sd(data.semana, data.dia, data.importe)
# print("saldo: {}".format(data["saldo diario"]))

# organizar columnas conforme a la lista nueva
data = data[final_columns]

# TODO QUITAR: BORRAR MES0
data.drop(["mes"], axis=1, inplace=True)

# guardar nuevo archivo sin columna de index
data.to_excel("outputs/viaticos2020.xlsx", index=False)
print("\nDataframe final:")
print(data.head())

# TODO: MANIPULAR EXCEL
wb = openpyxl.load_workbook("outputs/viaticos2020.xlsx")
ws = wb.active
cells = ws["A1:L00"]

# tamaño de celda
# ws.column_dimensions['A'].auto_size = True
# ws.column_dimensions['A'].bestFit = True
# ws.column_dimensions['A'].width = 15
# ws.row_dimensions[1].height = 30

# aplicar formatos
aplicarFormatos()

# crearformula
crearFormula()

# ajustar ancho
ajustarAncho()

# crear el nuevo archivo
wb.save("outputs/viaticos_formato.xlsx")

# TODO LLENAR PDF
# cargar archivo xls en el directorio
excel = "outputs/viaticos2020.xlsx"
# dataframe completo, si no hay fecha no es una columna valida
data = pd.read_excel(excel, header=0, names=columns_excel)
data = data.dropna(subset=["fecha"])
print("\nARCHIVO FINAL ORDENADO:\n")
print(data.head())
# llenar el nuevo dataframe
reporte = og(data)
